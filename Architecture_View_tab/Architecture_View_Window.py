# Architecture_View_Window.py
import os
import sys
import csv
from PyQt5.QtCore import QTimer, Qt, QSize, QRect
from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QTreeWidget,
    QTreeWidgetItem,
    QPushButton,
    QMessageBox,
    QHeaderView,
    QFileDialog,
    QApplication,
    QStyleFactory,
    QGroupBox,
    QSizePolicy,
    QStyledItemDelegate,
    QStyle,
    QLabel,
)
from PyQt5.QtGui import QColor, QIcon, QPixmap, QFont
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from database import (
    get_connection,
    init_db,
    get_current_project_id,
    get_module_subsystem_id,
    get_connector_subsystem_id,
    get_pin_subsystem_id,
    delete_subsystem_guarded,
    UnauthorizedError,
)
from auth_manager import auth
from access_control import can_edit_subsystem
from Architecture_View_tab.Module_Dialog import ModuleDialog
from Architecture_View_tab.Connector_Dialog import ConnectorDialog
from Architecture_View_tab.Pin_Dialog import PinDialog

# Import new style system
from styles.style_manager import (
    style_manager,
    register_widget,
    create_styled_button,
    auto_style_widget,
)
from styles.design_system import Colors, Typography, Spacing, BorderRadius
from styles.theme_manager import theme_manager

ITEM_TYPE_SUBSYSTEM = 1
ITEM_TYPE_MODULE = 2
ITEM_TYPE_CONNECTOR = 3
ITEM_TYPE_PIN = 4


class ModernTreeWidget(QTreeWidget):
    """درخت مدرن با پشتیبانی از تم‌های مختلف"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.apply_theme_style()

        # اتصال به تغییر تم
        style_manager.theme_changed.connect(self.apply_theme_style)

    def apply_theme_style(self):
        """اعمال استایل بر اساس تم فعلی"""
        style = f"""
            QTreeWidget {{
                background: {theme_manager.get_color('primary_dark')};
                color: {theme_manager.get_color('text_primary')};
                font-family: {Typography.FONT_FAMILY};
                font-size: {Typography.SIZE_MEDIUM};
                font-weight: {Typography.WEIGHT_MEDIUM};
                border: 1px solid {theme_manager.get_color('primary_light')};
                outline: none;
                selection-background-color: rgba(74, 144, 226, 40);
                show-decoration-selected: 1;
                alternate-background-color: rgba(62, 66, 99, 20);
            }}

            QTreeWidget QHeaderView::section {{
                background: {theme_manager.get_gradient("primary", "x1:0, y1:0, x2:1, y2:0")};
                color: white;
                font-family: {Typography.FONT_FAMILY};
                font-size: {Typography.SIZE_LARGE};
                font-weight: {Typography.WEIGHT_BOLD};
                padding: {Spacing.MD};
                border: 1px solid {theme_manager.get_color('primary_light')};
                border-radius: {BorderRadius.SMALL};
            }}

            QTreeWidget QScrollBar:vertical, QTreeWidget QScrollBar:horizontal {{
                width: 0px;
                height: 0px;
                background: transparent;
            }}
            
            QTreeWidget::item {{
                background: transparent;
                border: 1px solid rgba(74, 144, 226, 40);
                padding: {Spacing.LG} {Spacing.SM};
                margin: 1px 0px;
                border-radius: {BorderRadius.MEDIUM};
                min-height: 20px;
            }}
            
            QTreeWidget::item:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(74, 144, 226, 40), 
                    stop:1 rgba(155, 89, 182, 25));
                border: 1px solid rgba(74, 144, 226, 120);
                color: #ffffff;
            }}
            
            QTreeWidget::item:selected {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(74, 144, 226, 70), 
                    stop:1 rgba(155, 89, 182, 40));
                border: 1px solid rgba(74, 144, 226, 150);
                color: #ffffff;
                font-weight: 600;
            }}
            
            QTreeWidget::indicator {{
                width: 18px;
                height: 18px;
                border-radius: 9px;
                border: 2px solid #bdc3c7;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #ecf0f1);
                margin: 2px;
            }}
            
            QTreeWidget::indicator:unchecked {{
                border: 2px solid #95a5a6;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #ecf0f1);
            }}
            
            QTreeWidget::indicator:unchecked:hover {{
                border: 2px solid #3498db;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ffffff, stop:1 #e8f4fd);
            }}
            
            QTreeWidget::indicator:checked {{
                border: 2px solid #27ae60;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2ecc71, stop:1 #27ae60);
            }}
            
            QTreeWidget::indicator:checked:hover {{
                border: 2px solid #229954;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #58d68d, stop:1 #2ecc71);
            }}
            
            QTreeWidget::indicator:indeterminate {{
                border: 2px solid #f39c12;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #f1c40f, stop:1 #f39c12);
            }}
            
            QTreeWidget::branch {{
                background: transparent;
                border: none;
            }}
            
            QTreeWidget::branch:has-siblings:!adjoins-item,
            QTreeWidget::branch:has-siblings:adjoins-item,
            QTreeWidget::branch:!has-children:!has-siblings:adjoins-item {{
                border: none;
                background: transparent;
            }}
            
            QTreeWidget::branch:has-children:!has-siblings:closed,
            QTreeWidget::branch:closed:has-children:has-siblings {{
                border: none;
                background: transparent;
            }}
            
            QTreeWidget::branch:open:has-children:!has-siblings,
            QTreeWidget::branch:open:has-children:has-siblings {{
                border: none;
                background: transparent;
            }}
            
            QTreeWidget::branch:has-children:!has-siblings:closed:hover,
            QTreeWidget::branch:closed:has-children:has-siblings:hover {{
                background: rgba(52, 152, 219, 30);
                border-radius: 4px;
            }}
        """
        self.setStyleSheet(style)


class CustomTreeWidget(ModernTreeWidget):

    def mouseDoubleClickEvent(self, event):
        item = self.itemAt(event.pos())
        if item:
            self.itemDoubleClicked.emit(item, self.columnAt(event.pos().x()))
        else:
            super().mouseDoubleClickEvent(event)


class CenteredIconDelegate(QStyledItemDelegate):

    def paint(self, painter, option, index):
        if option.state & QStyle.State_Selected:
            painter.fillRect(option.rect, option.palette.highlight())

        if index.column() == 4:
            icon = index.data(Qt.DecorationRole)
            if icon and not icon.isNull():
                icon_size = option.decorationSize
                if icon_size.width() == 0 or icon_size.height() == 0:
                    icon_size = QSize(16, 16)
                rect = option.rect
                x = rect.x() + (rect.width() - icon_size.width()) / 2
                y = rect.y() + (rect.height() - icon_size.height()) / 2
                icon_rect = QRect(int(x), int(y), icon_size.width(), icon_size.height())
                icon.paint(painter, icon_rect, Qt.AlignCenter)
                return
        if index.column() != 0:
            option.displayAlignment = Qt.AlignCenter
        super().paint(painter, option, index)


class ArchitectureViewTab(QWidget):
    """تب نمای معماری با سیستم استایل جدید"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._checked = []  # List of checked items
        self._color_icon_cache = {}
        self.setObjectName("ArchitectureViewTab")
        self.setAttribute(Qt.WA_StyledBackground, True)

        # اعمال استایل پایه
        self.apply_main_widget_style()

        # اتصال به تغییر تم
        style_manager.theme_changed.connect(self.on_theme_changed)

        self.init_ui()
        self._expanded = False
        self._all_selected = False
        self.load_data_tree()
        self.data_tree.itemChanged.connect(self._on_item_changed)
        self.update_button_states()
        auth.auth_changed.connect(self.apply_access_policy)
        self.apply_access_policy()

    def update_delete_button(self):
        self.delete_item_btn.setEnabled(self._can_delete_checked())

    def _on_item_changed(self, item, column):
        if item.checkState(0) == Qt.Checked:
            if item not in self._checked:
                self._checked.append(item)
        else:
            if item in self._checked:
                self._checked.remove(item)
        QTimer.singleShot(0, self.update_delete_button)

    def apply_main_widget_style(self):
        """اعمال استایل به ویجت اصلی"""
        style = f"""
            QWidget#ArchitectureViewTab {{
                background: {theme_manager.get_color('primary_dark')};
                border: 1px solid {theme_manager.get_color('primary_light')};
                border-radius: {BorderRadius.XLARGE};
            }}
            
            QMessageBox {{
                background: {theme_manager.get_color('primary_dark')}; 
                color: {theme_manager.get_color('text_primary')};    
                font-family: {Typography.FONT_FAMILY};
                font-size: {Typography.SIZE_MEDIUM};
            }}

            QMessageBox QLabel {{
                color: {theme_manager.get_color('text_primary')};   
            }}

            QMessageBox QPushButton {{
                background: {theme_manager.get_gradient("primary")};
                border: 1px solid {theme_manager.get_color('accent')};
                border-radius: {BorderRadius.LARGE};
                color: {theme_manager.get_color('text_primary')};
                font-family: {Typography.FONT_FAMILY};
                font-size: {Typography.SIZE_MEDIUM};
                font-weight: {Typography.WEIGHT_BOLD};
                padding: {Spacing.LG} {Spacing.XL};
            }}
                           
            QComboBox QAbstractItemView {{
                background: {theme_manager.get_color('primary_dark')};
                color: {theme_manager.get_color('text_primary')};  
                selection-background-color: rgba(74, 144, 226, 40);
                selection-color: #ffffff;
                border: 1px solid {theme_manager.get_color('primary_light')};
                border-radius: {BorderRadius.MEDIUM};
                padding: {Spacing.MD};
            }}

            QComboBox QAbstractItemView::item {{
                padding: {Spacing.LG} {Spacing.XL};
            }}

            QComboBox QAbstractItemView::item:hover {{
                background-color: rgba(74, 144, 226, 30);
                color: white;
            }}
        """
        self.setStyleSheet(style)

    def on_theme_changed(self, theme_name):
        """هندل تغییر تم"""
        self.apply_main_widget_style()
        self.update_header_style()
        self.update_group_styles()

    def update_header_style(self):
        """به‌روزرسانی استایل هدر"""
        if hasattr(self, "header_widget"):
            header_style = f"""
                QWidget {{
                    background: {theme_manager.get_gradient("primary", "x1:0, y1:0, x2:1, y2:0")};
                    border-radius: {BorderRadius.LARGE};
                    border: 1px solid {theme_manager.get_color('primary_light')};
                }}
                QLabel {{
                    color: {theme_manager.get_color('text_primary')};
                    font-family: {Typography.FONT_FAMILY};
                    font-size: {Typography.SIZE_XLARGE};
                    font-weight: {Typography.WEIGHT_BOLD};
                    background: transparent;
                    border: none;
                }}
            """
            self.header_widget.setStyleSheet(header_style)

    def update_group_styles(self):
        """به‌روزرسانی استایل گروه‌ها"""
        group_style = f"""
            QGroupBox {{
                background: {theme_manager.get_gradient("primary", "x1:0, y1:0, x2:0, y2:1")};
                border: 1px solid {theme_manager.get_color('primary_light')};
                border-radius: {BorderRadius.LARGE};
                font-family: {Typography.FONT_FAMILY};
                font-size: {Typography.SIZE_MEDIUM};
                font-weight: {Typography.WEIGHT_BOLD};
                color: {theme_manager.get_color('text_primary')};
                margin-top: 10px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 3px;
                background: transparent;
            }}
        """

        # اعمال به همه گروه‌ها
        for widget in self.findChildren(QGroupBox):
            widget.setStyleSheet(group_style)

    def init_ui(self):
        """راه‌اندازی رابط کاربری"""
        font = QFont("Roboto Mono", 13, QFont.Medium)
        self.setFont(font)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)

        # Header widget
        self.header_widget = QWidget()
        self.header_widget.setFixedHeight(35)
        self.update_header_style()

        header_layout = QHBoxLayout(self.header_widget)
        header_layout.setContentsMargins(12, 0, 12, 0)
        header_label = QLabel("🛰️ Satellite Architecture")
        header_label.setFont(QFont("Roboto Mono", 16, QFont.Bold))
        header_layout.addWidget(header_label)
        header_layout.addStretch()
        main_layout.addWidget(self.header_widget)

        # Controls Layout
        top_layout = QHBoxLayout()

        # Left panel: Module, Connector, Pin groups
        left_panel = QHBoxLayout()
        self.create_control_groups(left_panel)
        left_panel.addStretch()
        top_layout.addLayout(left_panel, 3)

        # Right panel: Export, Controls
        right_panel = QHBoxLayout()
        self.create_export_controls(right_panel)
        right_panel.addStretch()
        top_layout.addLayout(right_panel, -2)

        main_layout.addLayout(top_layout)

        # Tree widget
        self.create_tree_widget(main_layout, font)

        # Connect all signals
        self.connect_signals()

    def create_control_groups(self, layout):
        """ایجاد گروه‌های کنترلی"""
        for add_text, edit_text in [
            ("Add Module", "Edit Module"),
            ("Add Connector", "Edit Connector"),
            ("Add Pin", "Edit Pin"),
        ]:
            group = QGroupBox("")
            vbox = QVBoxLayout()

            title_lower = add_text.lower().replace("add ", "")

            # ایجاد دکمه‌ها با سیستم جدید
            add_btn = create_styled_button(f"➕ {add_text}", "normal")
            edit_btn = create_styled_button(f"✏️ {edit_text}", "normal")

            setattr(self, f"add_{title_lower}_btn", add_btn)
            setattr(self, f"edit_{title_lower}_btn", edit_btn)

            vbox.addWidget(add_btn)
            vbox.addWidget(edit_btn)
            group.setLayout(vbox)
            layout.addWidget(group)

    def create_export_controls(self, layout):
        """ایجاد کنترل‌های صادرات"""
        # Export group
        export_group = QGroupBox("")
        export_vbox = QVBoxLayout()

        self.export_excel_btn = create_styled_button("📊 Export Excel", "normal")
        self.export_csv_btn = create_styled_button("📄 Export CSV", "normal")

        export_vbox.addWidget(self.export_excel_btn)
        export_vbox.addWidget(self.export_csv_btn)
        export_group.setLayout(export_vbox)
        layout.addWidget(export_group)

        # Controls group
        controls_group = QGroupBox("")
        controls_vbox = QVBoxLayout()

        self.toggle_expand_btn = create_styled_button("📂 Expand All", "normal")
        self.select_toggle_btn = create_styled_button("✅ Select All", "normal")
        self.delete_item_btn = create_styled_button("🗑️ Delete Selected", "normal")

        controls_vbox.addWidget(self.toggle_expand_btn)
        controls_vbox.addWidget(self.select_toggle_btn)
        controls_vbox.addWidget(self.delete_item_btn)
        controls_group.setLayout(controls_vbox)
        layout.addWidget(controls_group)

        # اعمال استایل گروه‌ها
        self.update_group_styles()

    def create_tree_widget(self, layout, font):
        """ایجاد ویجت درختی"""
        self.data_tree = CustomTreeWidget()
        self.data_tree.setFont(font)
        self.data_tree.setHeaderLabels(
            [
                "Name",
                "Type",
                "Mass(kg)",
                "Power(mW)",
                "Color",
                "Num Conn.",
                "Pin #",
                "Pin Type",
                "Voltage(V)",
                "Current(mA)",
            ]
        )
        self.data_tree.setColumnCount(10)

        # تنظیمات هدر
        header = self.data_tree.header()
        # ستون اول دستی تنظیم می‌شود
        header.setSectionResizeMode(0, QHeaderView.Interactive)
        # همه ستون‌های دیگر ثابت و برابر
        for col in range(1, self.data_tree.columnCount()):
            header.setSectionResizeMode(col, QHeaderView.Fixed)

        # تنظیمات درخت
        self.data_tree.setAlternatingRowColors(True)
        self.data_tree.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.data_tree.setRootIsDecorated(True)
        self.data_tree.setAnimated(True)
        self.data_tree.setIndentation(25)

        # Delegate برای آیکون‌ها
        delegate = CenteredIconDelegate(self.data_tree)
        for col in range(self.data_tree.columnCount()):
            self.data_tree.headerItem().setTextAlignment(col, Qt.AlignCenter)
            self.data_tree.setItemDelegateForColumn(col, delegate)

        layout.addWidget(self.data_tree)

        # تنظیم سایز ستون‌ها پس از اضافه شدن به layout
        self.data_tree.resizeEvent = self.on_tree_resize

    def on_tree_resize(self, event):
        """تنظیم مجدد سایز ستون‌ها هنگام تغییر سایز درخت"""
        super(CustomTreeWidget, self.data_tree).resizeEvent(event)
        self.resize_columns()

    def calculate_first_column_width(self):
        """محاسبه عرض مناسب برای ستون اول بر اساس محتوا"""
        return 400

    def resize_columns(self):
        """تنظیم سایز همه ستون‌ها"""
        if not hasattr(self, "data_tree") or self.data_tree.columnCount() == 0:
            return

        header = self.data_tree.header()

        # محاسبه و تنظیم ستون اول بر اساس محتوا
        first_column_width = self.calculate_first_column_width()
        header.resizeSection(0, first_column_width)

        # محاسبه عرض در دسترس برای ستون‌های باقی مانده
        total_width = self.data_tree.viewport().width()
        remaining_width = total_width - first_column_width

        # تعداد ستون‌های باقی مانده
        remaining_columns = self.data_tree.columnCount() - 1

        if remaining_columns > 0:
            # عرض هر ستون (برابر برای همه)
            column_width = max(
                80, remaining_width // remaining_columns
            )  # حداقل 80 پیکسل

            # تنظیم سایز همه ستون‌های باقی مانده
            for col in range(1, self.data_tree.columnCount()):
                header.resizeSection(col, column_width)

    def resize_first_column(self):
        """تنظیم اندازه ستون اول و بقیه ستون‌ها"""
        self.resize_columns()

    def connect_signals(self):
        """اتصال سیگنال‌ها"""
        # Module signals
        self.add_module_btn.clicked.connect(self.handle_add_module)
        self.edit_module_btn.clicked.connect(self.handle_edit_module)

        # Connector signals
        self.add_connector_btn.clicked.connect(self.handle_add_connector)
        self.edit_connector_btn.clicked.connect(self.handle_edit_connector)

        # Pin signals
        self.add_pin_btn.clicked.connect(self.handle_add_pin)
        self.edit_pin_btn.clicked.connect(self.handle_edit_pin)

        # Control signals
        self.delete_item_btn.clicked.connect(self.handle_delete_item)
        self.select_toggle_btn.clicked.connect(self.handle_select_toggle)

        # Export signals
        self.export_csv_btn.clicked.connect(self.handle_export_csv)
        self.export_excel_btn.clicked.connect(self.handle_export_excel)

        # Tree signals
        self.data_tree.itemSelectionChanged.connect(self.update_button_states)
        self.data_tree.itemChanged.connect(self.update_button_states)
        self.data_tree.itemDoubleClicked.connect(self.handle_item_double_clicked)
        self.data_tree.itemChanged.connect(self.handle_item_check_change)

        # Expand signals
        self.toggle_expand_btn.clicked.connect(self.handle_toggle_expand)
        self.data_tree.itemExpanded.connect(self.update_expand_button)
        self.data_tree.itemCollapsed.connect(self.update_expand_button)
        self.data_tree.itemChanged.connect(self.update_selection_button)

    def create_color_icon(self, hex_color):
        """ایجاد آیکون رنگی"""
        pix = QPixmap(16, 16)
        pix.fill(QColor(hex_color))
        return QIcon(pix)

    def _get_color_icon(self, hex_color):
        """Reuse color icons to avoid repeatedly creating pixmaps."""
        color = hex_color or "#C8C8FF"
        if color not in self._color_icon_cache:
            self._color_icon_cache[color] = self.create_color_icon(color)
        return self._color_icon_cache[color]

    def load_data_tree(self):
        """Load tree data from PostgreSQL for the current project."""
        self.data_tree.clear()

        pid = get_current_project_id()
        if pid is None:
            QMessageBox.warning(
                self, "No Project", "Please open or create a project first."
            )
            return

        # different fonts for levels
        subsystem_font = QFont("Roboto Mono", 14, QFont.Bold)
        module_font = QFont("Roboto Mono", 13, QFont.Bold)
        connector_font = QFont("Roboto Mono", 12, QFont.Bold)
        pin_font = QFont("Roboto Mono", 11, QFont.Bold)

        self.data_tree.setUpdatesEnabled(False)
        try:
            with get_connection() as conn:
                cur = conn.cursor()

                cur.execute(
                    "SELECT id, name FROM subsystems WHERE project_id = %s ORDER BY name",
                    (pid,),
                )
                subsystem_rows = cur.fetchall()

                cur.execute(
                    "SELECT id, name, subsystem_id, mass, power, num_connectors, color "
                    "FROM modules WHERE project_id = %s ORDER BY subsystem_id, name",
                    (pid,),
                )
                module_rows = cur.fetchall()

                cur.execute(
                    "SELECT id, name, module_id, number_of_pins, color "
                    "FROM connectors WHERE project_id = %s ORDER BY module_id, name",
                    (pid,),
                )
                connector_rows = cur.fetchall()

                cur.execute(
                    "SELECT id, name, pin_number, pin_type, is_ground, value, current, connector_id "
                    "FROM pins WHERE project_id = %s ORDER BY connector_id, pin_number",
                    (pid,),
                )
                pin_rows = cur.fetchall()

                modules_by_subsystem = {}
                for row in module_rows:
                    mod_id, _, subsystem_id, *_ = row
                    modules_by_subsystem.setdefault(subsystem_id, []).append(row)

                connectors_by_module = {}
                for connector_row in connector_rows:
                    conn_id, _, module_id, *_ = connector_row
                    connectors_by_module.setdefault(module_id, []).append(connector_row)

                pins_by_connector = {}
                for pin_row in pin_rows:
                    _, _, _, _, _, _, _, connector_id = pin_row
                    pins_by_connector.setdefault(connector_id, []).append(pin_row)

                for sub_id, sub_name in subsystem_rows:
                    sub_item = QTreeWidgetItem(
                        self.data_tree,
                        [f"◻️ {sub_name}", "Subsystem", "", "", "", "", "", "", "", ""],
                    )
                    sub_item.setFont(0, subsystem_font)
                    sub_item.setData(0, Qt.UserRole, sub_id)
                    sub_item.setData(0, Qt.UserRole + 1, ITEM_TYPE_SUBSYSTEM)
                    sub_item.setForeground(
                        0, QColor(theme_manager.get_color("text_primary"))
                    )

                    for mod_index, (
                        mod_id,
                        name,
                        _,
                        mass,
                        power,
                        nconn,
                        color,
                    ) in enumerate(modules_by_subsystem.get(sub_id, []), start=1):
                        display_name = f"{mod_index}. {name}"
                        mod_item = QTreeWidgetItem(
                            sub_item,
                            [
                                f"⚙️ {display_name}",
                                "Module",
                                f"{(mass or 0):.3f}",
                                f"{(power or 0):.2f}",
                                "",
                                str(nconn or 0),
                                "",
                                "",
                                "",
                                "",
                            ],
                        )
                        mod_item.setFont(0, module_font)
                        mod_item.setFlags(mod_item.flags() | Qt.ItemIsUserCheckable)
                        mod_item.setCheckState(0, Qt.Unchecked)
                        mod_item.setIcon(4, self._get_color_icon(color or "#C8C8FF"))
                        mod_item.setData(0, Qt.UserRole, mod_id)
                        mod_item.setData(0, Qt.UserRole + 1, ITEM_TYPE_MODULE)
                        mod_item.setForeground(
                            0, QColor(theme_manager.get_color("text_secondary"))
                        )

                        for conn_index, (
                            cid,
                            cname,
                            _,
                            total_pins,
                            ccolor,
                        ) in enumerate(connectors_by_module.get(mod_id, []), start=1):
                            pin_rows_for_connector = pins_by_connector.get(cid, [])
                            used = len(pin_rows_for_connector)

                            display_cname = f"{conn_index}. {cname}"
                            conn_item = QTreeWidgetItem(
                                mod_item,
                                [
                                    f"🔌 {display_cname}",
                                    "Connector",
                                    "",
                                    "",
                                    "",
                                    "",
                                    f"{used}/{total_pins or 0}",
                                    "",
                                    "",
                                    "",
                                ],
                            )
                            conn_item.setFont(0, connector_font)
                            conn_item.setFlags(
                                conn_item.flags() | Qt.ItemIsUserCheckable
                            )
                            conn_item.setCheckState(0, Qt.Unchecked)
                            conn_item.setIcon(
                                4, self._get_color_icon(ccolor or "#C8C8FF")
                            )
                            conn_item.setData(0, Qt.UserRole, cid)
                            conn_item.setData(0, Qt.UserRole + 1, ITEM_TYPE_CONNECTOR)
                            conn_item.setForeground(0, QColor("#7f8c8d"))

                            for pin_index, (
                                pid_row,
                                pname,
                                pnum,
                                ptype,
                                isg,
                                val,
                                curm,
                                _,
                            ) in enumerate(pin_rows_for_connector, start=1):
                                if (ptype or "").lower() == "voltage":
                                    tstr = "GND" if isg else "VCC"
                                    vstr = "" if isg else (f"{(val or 0):.2f}")
                                else:
                                    tstr, vstr = (ptype or ""), ""
                                curr_text = (
                                    f"{(curm or 0):.2f}" if curm is not None else ""
                                )
                                display_pname = f"{pin_index}. {pname}"
                                pin_item = QTreeWidgetItem(
                                    conn_item,
                                    [
                                        f"📍 {display_pname}",
                                        "Pin",
                                        "",
                                        "",
                                        "",
                                        "",
                                        str(pnum if pnum is not None else ""),
                                        tstr,
                                        vstr,
                                        curr_text,
                                    ],
                                )
                                pin_item.setFont(0, pin_font)
                                pin_item.setFlags(
                                    pin_item.flags() | Qt.ItemIsUserCheckable
                                )
                                pin_item.setCheckState(0, Qt.Unchecked)
                                pin_item.setData(0, Qt.UserRole, pid_row)
                                pin_item.setData(0, Qt.UserRole + 1, ITEM_TYPE_PIN)
                                pin_item.setForeground(0, QColor("#95a5a6"))

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load data:\n{e}")
        finally:
            self.data_tree.setUpdatesEnabled(True)
            self.data_tree.expandAll()
            self.resize_columns()

    def _can_delete_checked(self):
        any_checked = bool(self.checked_items())
        if not any_checked:
            return False

        sids = []
        types = []

        for item in self.checked_items():
            try:
                _ = item.text(0)
            except RuntimeError:
                continue

            # walk up to find the parent subsystem
            parent = item
            while parent.parent() is not None:
                parent = parent.parent()

            # now parent is top-level subsystem
            t = item.data(0, Qt.UserRole + 1)
            iid = item.data(0, Qt.UserRole)
            sid = self._resolve_subsystem_id(t, iid)  # optional extra param
            sids.append(sid)
            types.append(t)

        if not all(can_edit_subsystem(s) for s in sids):
            return False

        type_to_perm = {
            ITEM_TYPE_MODULE: "module.delete",
            ITEM_TYPE_CONNECTOR: "connector.delete",
            ITEM_TYPE_PIN: "pin.delete",
            ITEM_TYPE_SUBSYSTEM: "subsystem.delete",
        }
        required = set(type_to_perm.get(t, "") for t in types) - {""}
        return all(auth.has_perm(p) for p in required)

    def _resolve_subsystem_id(self, item_type, item_id):
        if item_type is None or item_id is None:
            return None
        if item_type == ITEM_TYPE_SUBSYSTEM:
            return item_id
        if item_type == ITEM_TYPE_MODULE:
            return get_module_subsystem_id(item_id)
        if item_type == ITEM_TYPE_CONNECTOR:
            return get_connector_subsystem_id(item_id)
        if item_type == ITEM_TYPE_PIN:
            return get_pin_subsystem_id(item_id)
        return None

    def update_button_states(self):
        sel = self.data_tree.selectedItems()
        item_type = sel[0].data(0, Qt.UserRole + 1) if sel else None
        item_id = sel[0].data(0, Qt.UserRole) if sel else None

        # figure selected subsystem for scope checks
        selected_sid = self._resolve_subsystem_id(item_type, item_id)

        # helpers
        def _en(name, ok):
            btn = getattr(self, name, None)
            if btn is not None:
                btn.setEnabled(bool(ok))

        # permissions
        can_mod_create = auth.has_perm("module.create") and can_edit_subsystem(
            selected_sid
        )
        can_mod_edit = (
            auth.has_perm("module.edit")
            and (item_type == ITEM_TYPE_MODULE)
            and can_edit_subsystem(selected_sid)
        )

        can_con_create = (
            auth.has_perm("connector.create")
            and (item_type == ITEM_TYPE_MODULE)
            and can_edit_subsystem(selected_sid)
        )
        can_con_edit = (
            auth.has_perm("connector.edit")
            and (item_type == ITEM_TYPE_CONNECTOR)
            and can_edit_subsystem(selected_sid)
        )

        can_pin_create = (
            auth.has_perm("pin.create")
            and (item_type == ITEM_TYPE_CONNECTOR)
            and can_edit_subsystem(selected_sid)
        )
        can_pin_edit = (
            auth.has_perm("pin.edit")
            and (item_type == ITEM_TYPE_PIN)
            and can_edit_subsystem(selected_sid)
        )

        # map to our button attributes created in create_control_groups
        _en("add_module_btn", can_mod_create or auth.is_system())
        _en("edit_module_btn", can_mod_edit or auth.is_system())

        _en("add_connector_btn", can_con_create or auth.is_system())
        _en("edit_connector_btn", can_con_edit or auth.is_system())

        _en("add_pin_btn", can_pin_create or auth.is_system())
        _en("edit_pin_btn", can_pin_edit or auth.is_system())

        # delete enabled only if ALL checked items are within scope and user has proper delete perms

        self.delete_item_btn.setEnabled(self._can_delete_checked())

        # export/expand/select are always allowed
        self.export_excel_btn.setEnabled(True)
        self.export_csv_btn.setEnabled(True)
        self.toggle_expand_btn.setEnabled(True)
        self.select_toggle_btn.setEnabled(True)

    def checked_items(self):
        return list(self._checked)

    def _has_checked_ancestor(self, item, checked_items):
        parent = item.parent()
        while parent is not None:
            if parent in checked_items:
                return True
            parent = parent.parent()
        return False

    def handle_select_toggle(self):
        """تغییر وضعیت انتخاب همه"""
        state = Qt.Checked if not self._all_selected else Qt.Unchecked

        # Block signals during batch operation to avoid O(n²) updates
        self.data_tree.blockSignals(True)

        for i in range(self.data_tree.topLevelItemCount()):
            item = self.data_tree.topLevelItem(i)
            self._set_all_children_checked(item, state)

        self.data_tree.blockSignals(False)

        # Rebuild _checked list once instead of during each signal
        self._rebuild_checked_list()
        self._all_selected = not self._all_selected
        self.update_selection_button()
        self.resize_first_column()
        self.update_delete_button()

    def _rebuild_checked_list(self):
        """Rebuild _checked list by scanning entire tree once (O(n))"""
        self._checked.clear()
        stack = [
            self.data_tree.topLevelItem(i)
            for i in range(self.data_tree.topLevelItemCount())
        ]
        while stack:
            item = stack.pop()
            if item.checkState(0) == Qt.Checked:
                self._checked.append(item)
            for i in range(item.childCount()):
                stack.append(item.child(i))

    def _set_all_children_checked(self, parent_item, state):
        """تنظیم وضعیت چک برای همه فرزندان"""
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            child.setCheckState(0, state)
            self._set_all_children_checked(child, state)

    def update_selection_button(self):
        """به‌روزرسانی متن دکمه انتخاب"""
        if self.is_all_selected():
            self.select_toggle_btn.setText("🟩 Deselect All")
            self.select_toggle_btn.setToolTip("Deselect All")
        else:
            self.select_toggle_btn.setText("✅ Select All")
            self.select_toggle_btn.setToolTip("Select All")

    def is_all_selected(self):
        """بررسی انتخاب همه آیتم‌ها"""

        def check_children(item):
            for i in range(item.childCount()):
                child = item.child(i)
                if child.checkState(0) != Qt.Checked:
                    return False
                if not check_children(child):
                    return False
            return True

        for i in range(self.data_tree.topLevelItemCount()):
            item = self.data_tree.topLevelItem(i)
            if not check_children(item):
                return False
        return True

    def update_expand_button(self):
        """به‌روزرسانی دکمه گسترش"""
        if self.is_all_expanded():
            self.toggle_expand_btn.setText("📁 Collapse All")
            self.toggle_expand_btn.setToolTip("Collapse All")
        else:
            self.toggle_expand_btn.setText("📂 Expand All")
            self.toggle_expand_btn.setToolTip("Expand All")
        self.resize_first_column()

    def is_all_expanded(self):
        stack = [
            self.data_tree.topLevelItem(i)
            for i in range(self.data_tree.topLevelItemCount())
        ]

        while stack:
            item = stack.pop()

            if not item.isExpanded():
                return False

            for i in range(item.childCount()):
                stack.append(item.child(i))

        return True

    def get_selected_item(self):
        """دریافت آیتم انتخاب شده"""
        sel = self.data_tree.selectedItems()
        if not sel:
            return None, None
        it = sel[0]
        return it.data(0, Qt.UserRole), it.data(0, Qt.UserRole + 1)

    def handle_add_module(self):
        sid, t = self.get_selected_item()
        dlg = ModuleDialog(
            subsystem_id=(sid if t == ITEM_TYPE_SUBSYSTEM else None), parent=self
        )
        auto_style_widget(dlg)
        dlg.modules_updated.connect(self.load_data_tree)
        dlg.exec_()
        # dlg = ModuleDlg(sid if t == ITEM_TYPE_SUBSYSTEM else None, parent=self)
        # dlg.show()

    def handle_edit_module(self):
        mid, t = self.get_selected_item()
        if t != ITEM_TYPE_MODULE:
            return

        pid = get_current_project_id()
        if pid is None:
            return

        try:
            with get_connection() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT name, subsystem_id, mass, power, num_connectors, color "
                    "FROM modules WHERE id = %s AND project_id = %s",
                    (mid, pid),
                )
                row = cur.fetchone()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read module:\n{e}")
            return

        if not row:
            return

        data = {
            "id": mid,
            "name": row[0],
            "subsystem_id": row[1],
            "mass": row[2],
            "power": row[3],
            "num_connectors": row[4],
            "color": row[5],
        }
        dlg = ModuleDialog(subsystem_id=row[1], module_data=data, parent=self)
        auto_style_widget(dlg)
        dlg.modules_updated.connect(self.load_data_tree)
        dlg.exec_()

    def handle_add_connector(self):
        mid, t = self.get_selected_item()
        if t != ITEM_TYPE_MODULE:
            return

        dlg = ConnectorDialog(module_id=mid, parent=self)
        auto_style_widget(dlg)
        dlg.connectors_updated.connect(self.load_data_tree)
        dlg.exec_()

    def handle_edit_connector(self):
        cid, t = self.get_selected_item()
        if t != ITEM_TYPE_CONNECTOR:
            return

        pid = get_current_project_id()
        if pid is None:
            return

        try:
            with get_connection() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT name, module_id, number_of_pins, color "
                    "FROM connectors WHERE id = %s AND project_id = %s",
                    (cid, pid),
                )
                row = cur.fetchone()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read connector:\n{e}")
            return

        if not row:
            return

        data = {
            "id": cid,
            "name": row[0],
            "module_id": row[1],
            "number_of_pins": row[2],
            "color": row[3],
        }
        dlg = ConnectorDialog(module_id=row[1], connector_data=data, parent=self)
        auto_style_widget(dlg)
        dlg.connectors_updated.connect(self.load_data_tree)
        dlg.exec_()

    def handle_add_pin(self):
        cid, t = self.get_selected_item()
        if t != ITEM_TYPE_CONNECTOR:
            return

        dlg = PinDialog(connector_id=cid, parent=self)
        auto_style_widget(dlg)
        dlg.pins_updated.connect(self.load_data_tree)
        dlg.exec_()

    def handle_edit_pin(self):
        pid_row, t = self.get_selected_item()
        if t != ITEM_TYPE_PIN:
            return

        pid = get_current_project_id()
        if pid is None:
            return

        try:
            with get_connection() as conn:
                cur = conn.cursor()
                cur.execute(
                    "SELECT name, pin_number, pin_type, is_ground, value, current, description, connector_id "
                    "FROM pins WHERE id = %s AND project_id = %s",
                    (pid_row, pid),
                )
                row = cur.fetchone()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read pin:\n{e}")
            return

        if not row:
            return

        data = {
            "id": pid_row,
            "name": row[0],
            "pin_number": row[1],
            "pin_type": row[2],
            "is_ground": row[3],
            "value": row[4],
            "current": row[5],
            "description": row[6],
        }
        dlg = PinDialog(pin_data=data, connector_id=row[7], parent=self)
        auto_style_widget(dlg)
        dlg.pins_updated.connect(self.load_data_tree)
        dlg.exec_()

    def handle_delete_item(self):
        checked = self.checked_items()
        if not checked:
            QMessageBox.information(self, "Delete", "No items checked.")
            return

            # scope validation: ALL must be within user's editable subsystem(s)
        invalid = []
        for it in checked:
            try:
                _ = it.text(0)
            except RuntimeError:
                continue
            t = it.data(0, Qt.UserRole + 1)
            iid = it.data(0, Qt.UserRole)
            sid = self._resolve_subsystem_id(t, iid)
            if t == ITEM_TYPE_SUBSYSTEM:
                # subsystem removal is a system-admin-only operation
                if not auth.is_system():
                    invalid.append(it.text(0))
            elif not (auth.is_system() or can_edit_subsystem(sid)):
                invalid.append(it.text(0))
        if invalid:
            QMessageBox.warning(
                self,
                "Access denied",
                "You cannot delete items outside your subsystem.\n"
                "Invalid selection(s):\n- " + "\n- ".join(invalid),
            )
            return

        if (
            QMessageBox.question(
                self,
                "Confirm Delete",
                f"Delete {len(checked)} item(s)?",
                QMessageBox.Yes | QMessageBox.No,
            )
            != QMessageBox.Yes
        ):
            return

        pid = get_current_project_id()
        if pid is None:
            QMessageBox.warning(
                self, "No Project", "Please open or create a project first."
            )
            return

        subsystem_items = []
        try:
            items_to_delete = [
                it for it in checked if not self._has_checked_ancestor(it, checked)
            ]

            subsystem_items = [
                it for it in items_to_delete
                if it.data(0, Qt.UserRole + 1) == ITEM_TYPE_SUBSYSTEM
            ]
            regular_items = [
                it for it in items_to_delete
                if it.data(0, Qt.UserRole + 1) != ITEM_TYPE_SUBSYSTEM
            ]

            self.data_tree.blockSignals(True)
            # Regular module/connector/pin deletions
            with get_connection() as conn:
                cur = conn.cursor()
                for it in regular_items:
                    t = it.data(0, Qt.UserRole + 1)
                    iid = it.data(0, Qt.UserRole)

                    if t == ITEM_TYPE_PIN:
                        cur.execute(
                            "DELETE FROM pins WHERE id = %s AND project_id = %s",
                            (iid, pid),
                        )

                    elif t == ITEM_TYPE_CONNECTOR:
                        cur.execute(
                            "DELETE FROM connectors WHERE id = %s AND project_id = %s",
                            (iid, pid),
                        )

                    elif t == ITEM_TYPE_MODULE:
                        cur.execute(
                            "DELETE FROM modules WHERE id = %s AND project_id = %s",
                            (iid, pid),
                        )

                    parent = it.parent()
                    if parent is None:
                        top_index = self.data_tree.indexOfTopLevelItem(it)
                        if top_index >= 0:
                            self.data_tree.takeTopLevelItem(top_index)
                    else:
                        parent_index = parent.indexOfChild(it)
                        if parent_index >= 0:
                            parent.takeChild(parent_index)

                conn.commit()

            # Subsystem deletions: full cascade via the guarded DB helper
            for it in subsystem_items:
                sid = it.data(0, Qt.UserRole)
                try:
                    ok, msg = delete_subsystem_guarded(auth.user_id, sid)
                except UnauthorizedError:
                    QMessageBox.warning(
                        self,
                        "Access denied",
                        "Only the system admin can delete subsystems.",
                    )
                    continue
                if not ok:
                    QMessageBox.critical(
                        self, "Error", f"Failed to delete subsystem:\n{msg}"
                    )

            self.data_tree.blockSignals(False)
        except Exception as e:
            self.data_tree.blockSignals(False)
            QMessageBox.critical(self, "Error", f"Failed to delete items:\n{e}")
            return

        # Reload the tree so the UI matches the database exactly
        # (a subsystem deletion removes a whole subtree at once)
        if subsystem_items:
            self.load_data_tree()
        self._checked.clear()
        self._rebuild_checked_list()
        self.update_button_states()
        self.update_delete_button()
        self.update_selection_button()

    def handle_item_double_clicked(self, item, column):
        t = item.data(0, Qt.UserRole + 1)
        iid = item.data(0, Qt.UserRole)
        sid = self._resolve_subsystem_id(t, iid)

        # only allow drilling down if user can edit that subsystem (or is system)
        if not (auth.is_system() or can_edit_subsystem(sid)):
            QMessageBox.warning(
                self,
                "Access denied",
                "You don't have permission to modify this subsystem.",
            )
            return

        if t == ITEM_TYPE_SUBSYSTEM:
            self.handle_add_module()
        elif t == ITEM_TYPE_MODULE:
            self.handle_add_connector()
        elif t == ITEM_TYPE_CONNECTOR:
            self.handle_add_pin()

    def handle_item_check_change(self, item, column):
        """هندل تغییر چک باکس آیتم"""
        if column != 0 or not (item.flags() & Qt.ItemIsUserCheckable):
            return

        state = item.checkState(0)

        # Block signals during cascade to avoid O(n²) updates
        self.data_tree.blockSignals(True)

        stack = [item]
        while stack:
            c = stack.pop()
            for i in range(c.childCount()):
                child = c.child(i)
                child.setCheckState(0, state)
                stack.append(child)

        self.data_tree.blockSignals(False)

        # Rebuild checked list once after all changes
        self._rebuild_checked_list()
        self.update_selection_button()
        self.resize_first_column()
        self.update_delete_button()

    def handle_export_csv(self):
        """صادرات به CSV"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save CSV File", "", "CSV Files (*.csv)"
        )
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(
                    [
                        "Subsystem",
                        "Module",
                        "Module Mass(kg)",
                        "Module Power(mW)",
                        "Module Color",
                        "Module Num Connectors",
                        "Connector",
                        "Connector Color",
                        "Connector Total Pins",
                        "Pin Name",
                        "Pin Number",
                        "Pin Format/Type",
                        "Pin Voltage(V)",
                        "Pin Current(mA)",
                    ]
                )

                for i in range(self.data_tree.topLevelItemCount()):
                    subsys_item = self.data_tree.topLevelItem(i)
                    subsys = subsys_item.text(0).replace("🏢 ", "")

                    if subsys_item.childCount() == 0:
                        writer.writerow([subsys] + [""] * 13)
                        continue

                    for j in range(subsys_item.childCount()):
                        mod_item = subsys_item.child(j)
                        module = mod_item.text(0).replace("⚙️ ", "").split(". ", 1)[1]
                        m_mass = mod_item.text(2)
                        m_power = mod_item.text(3)
                        m_color = (
                            mod_item.icon(4)
                            .pixmap(16, 16)
                            .toImage()
                            .pixelColor(0, 0)
                            .name()
                        )
                        m_nconn = mod_item.text(5)

                        if mod_item.childCount() == 0:
                            writer.writerow(
                                [subsys, module, m_mass, m_power, m_color, m_nconn]
                                + [""] * 8
                            )
                            continue

                        for k in range(mod_item.childCount()):
                            conn_item = mod_item.child(k)
                            conn_name = (
                                conn_item.text(0).replace("🔌 ", "").split(". ", 1)[1]
                            )
                            conn_color = (
                                conn_item.icon(4)
                                .pixmap(16, 16)
                                .toImage()
                                .pixelColor(0, 0)
                                .name()
                            )
                            conn_pins = conn_item.text(6)

                            if conn_item.childCount() == 0:
                                writer.writerow(
                                    [
                                        subsys,
                                        module,
                                        m_mass,
                                        m_power,
                                        m_color,
                                        m_nconn,
                                        conn_name,
                                        conn_color,
                                        conn_pins,
                                    ]
                                    + [""] * 5
                                )
                                continue

                            for m in range(conn_item.childCount()):
                                pin_item = conn_item.child(m)
                                pin_name = (
                                    pin_item.text(0)
                                    .replace("📍 ", "")
                                    .split(". ", 1)[1]
                                )
                                writer.writerow(
                                    [
                                        subsys,
                                        module,
                                        m_mass,
                                        m_power,
                                        m_color,
                                        m_nconn,
                                        conn_name,
                                        conn_color,
                                        conn_pins,
                                        pin_name,
                                        pin_item.text(6),
                                        pin_item.text(7),
                                        pin_item.text(8),
                                        pin_item.text(9),
                                    ]
                                )

            QMessageBox.information(
                self, "Success", "Data exported to CSV successfully."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export CSV: {e}")

    def handle_export_excel(self):
        """صادرات به Excel"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Architecture"

        headers = [
            "Subsystem",
            "Module",
            "Module Mass(kg)",
            "Module Power(mW)",
            "Module Color",
            "Module Num Connectors",
            "Connector",
            "Connector Color",
            "Connector Total Pins",
            "Pin Name",
            "Pin Number",
            "Pin Format/Type",
            "Pin Voltage(V)",
            "Pin Current(mA)",
        ]
        ws.append(headers)

        rows_to_write = []
        for i in range(self.data_tree.topLevelItemCount()):
            subsys_item = self.data_tree.topLevelItem(i)
            subsys = subsys_item.text(0).replace("🏢 ", "")

            for j in range(subsys_item.childCount()):
                mod_item = subsys_item.child(j)
                module = mod_item.text(0).replace("⚙️ ", "").split(". ", 1)[1]
                m_mass = mod_item.text(2)
                m_power = mod_item.text(3)
                m_color = (
                    mod_item.icon(4).pixmap(16, 16).toImage().pixelColor(0, 0).name()
                )
                m_nconn = mod_item.text(5)

                for k in range(mod_item.childCount()):
                    conn_item = mod_item.child(k)
                    cname = conn_item.text(0).replace("🔌 ", "").split(". ", 1)[1]
                    ccolor = (
                        conn_item.icon(4)
                        .pixmap(16, 16)
                        .toImage()
                        .pixelColor(0, 0)
                        .name()
                    )
                    ctotal = conn_item.text(6)

                    if conn_item.childCount() == 0:
                        rows_to_write.append(
                            [
                                subsys,
                                module,
                                m_mass,
                                m_power,
                                m_color,
                                m_nconn,
                                cname,
                                ccolor,
                                ctotal,
                            ]
                            + [""] * 5
                        )
                    else:
                        for m in range(conn_item.childCount()):
                            pin = conn_item.child(m)
                            pin_name = pin.text(0).replace("📍 ", "").split(". ", 1)[1]
                            rows_to_write.append(
                                [
                                    subsys,
                                    module,
                                    m_mass,
                                    m_power,
                                    m_color,
                                    m_nconn,
                                    cname,
                                    ccolor,
                                    ctotal,
                                    pin_name,
                                    pin.text(6),
                                    pin.text(7),
                                    pin.text(8),
                                    pin.text(9),
                                ]
                            )

        for row in rows_to_write:
            ws.append(row)

        def merge_and_center(cols):
            for col in cols:
                start = 2
                prev = ws.cell(start, col).value
                for r in range(3, ws.max_row + 1):
                    val = ws.cell(r, col).value
                    if val != prev:
                        if r - 1 > start:
                            ws.merge_cells(
                                start_row=start,
                                start_column=col,
                                end_row=r - 1,
                                end_column=col,
                            )
                            ws.cell(start, col).alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                        prev = val
                        start = r
                if start < ws.max_row:
                    ws.merge_cells(
                        start_row=start,
                        start_column=col,
                        end_row=ws.max_row,
                        end_column=col,
                    )
                    ws.cell(start, col).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

        merge_and_center([1])
        merge_and_center([2, 3, 4, 5, 6])
        merge_and_center([7, 8, 9])

        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                length + 2
            )

        try:
            wb.save(file_path)
            QMessageBox.information(
                self, "Success", "Data exported to Excel successfully."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel: {e}")

    def handle_toggle_expand(self):
        """تغییر وضعیت گسترش همه آیتم‌ها"""
        if not self._expanded:
            self.data_tree.expandAll()
            self._expanded = True
        else:
            self.data_tree.collapseAll()
            self._expanded = False
        self.update_expand_button()

    def apply_access_policy(self):
        # Try to detect selected module id from your widgets; adjust as needed.
        selected_module_id = getattr(self, "selected_module_id", None)
        if selected_module_id is None and hasattr(self, "current_module_id"):
            selected_module_id = getattr(self, "current_module_id")
        subsystem_id = (
            get_module_subsystem_id(selected_module_id) if selected_module_id else None
        )

        # Helper to set enabled state if button exists
        def _set_enabled(name: str, enabled: bool):
            btn = getattr(self, name, None)
            if btn is not None:
                btn.setEnabled(enabled)

        # Create permissions (global)
        can_create = auth.has_perm("module.create")
        # Edit/Delete depend on both permission and scope
        can_edit = auth.has_perm("module.edit") and can_edit_subsystem(subsystem_id)
        can_delete = auth.has_perm("module.delete") and can_edit_subsystem(subsystem_id)

        # Adjust button names to match your actual attribute names:
        _set_enabled("btn_add_module", can_create)
        _set_enabled("btn_edit_module", can_edit)
        _set_enabled("btn_delete_module", can_delete)

        # You can similarly handle connectors/pins buttons if present:
        _set_enabled(
            "btn_add_connector",
            auth.has_perm("connector.create") and can_edit_subsystem(subsystem_id),
        )
        _set_enabled(
            "btn_edit_connector",
            auth.has_perm("connector.edit") and can_edit_subsystem(subsystem_id),
        )
        _set_enabled(
            "btn_delete_connector",
            auth.has_perm("connector.delete") and can_edit_subsystem(subsystem_id),
        )


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle(QStyleFactory.create("Fusion"))
    init_db()
    w = ArchitectureViewTab()
    w.show()
    sys.exit(app.exec_())
